from flask import Blueprint, request, jsonify, current_app, flash, redirect, url_for, render_template, send_from_directory
from app.models import Certifikat, Server
from app import db
from app.utils import allowed_file, is_valid_date
from datetime import datetime, timedelta
import pandas as pd
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from app import get_expiry_class  # Přidejte tento import na začátek souboru
from tempfile import NamedTemporaryFile
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

bp = Blueprint('certificates', __name__)

@bp.route('/pridat', methods=['GET', 'POST'])
def pridat_certifikat():
    if request.method == 'POST':
        try:
            expirace = datetime.strptime(request.form['expirace'], '%d.%m.%Y')
            
            novy_cert = Certifikat(
                server=request.form['server'],
                cesta=request.form['cesta'],
                nazev=request.form['nazev'],
                expirace=expirace,
                poznamka=request.form.get('poznamka', '')
            )
            db.session.add(novy_cert)
            db.session.commit()
            flash('Certifikát byl úspěšně přidán!')
            return redirect(url_for('main.index'))
        except Exception as e:
            current_app.logger.error(f'Chyba při přidávání certifikátu: {str(e)}')
            flash(f'Chyba při přidávání certifikátu: {str(e)}')
            return redirect(url_for('certificates.pridat_certifikat'))
    
    servery = Server.query.all()
    return render_template('formular.html', certifikat=None, servery=servery)

@bp.route('/get-edit-form/<int:id>')
def get_edit_form(id):
    try:
        certifikat = Certifikat.query.get_or_404(id)
        servery = Server.query.all()
        return render_template('edit_modal.html', certifikat=certifikat, servery=servery)
    except Exception as e:
        current_app.logger.error(f'Chyba při načítání editačního formuláře: {str(e)}')
        return jsonify({'error': str(e)}), 500

@bp.route('/upravit/<int:id>', methods=['GET', 'POST'])
def upravit_certifikat(id):
    certifikat = Certifikat.query.get_or_404(id)
    servery = Server.query.all()
    
    if request.method == 'POST':
        try:
            certifikat.server = request.form['server']
            certifikat.cesta = request.form['cesta']
            certifikat.nazev = request.form['nazev']
            
            # Zpracování data
            expirace_str = request.form['expirace']
            expirace = datetime.strptime(expirace_str, '%d.%m.%Y').date()
            certifikat.expirace = expirace
            
            certifikat.poznamka = request.form.get('poznamka', '')
            
            db.session.commit()
            flash('Certifikát byl úspěšně upraven', 'success')
            return redirect('/evidence_certifikatu')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'Chyba při úpravě certifikátu: {str(e)}')
            flash(f'Chyba při úpravě certifikátu: {str(e)}', 'error')
            return redirect('/evidence_certifikatu')
    
    return render_template('formular.html', 
                         certifikat=certifikat,
                         servery=servery)

@bp.route('/smazat/<int:id>')
def smazat_certifikat(id):
    try:
        certifikat = Certifikat.query.get_or_404(id)
        current_app.logger.info(f'Mazání certifikátu: {certifikat.nazev} ze serveru {certifikat.server}')
        db.session.delete(certifikat)
        db.session.commit()
        flash('Certifikát byl smazán!')
    except Exception as e:
        current_app.logger.error(f'Chyba při mazání certifikátu: {str(e)}')
        flash(f'Chyba při mazání certifikátu: {str(e)}')
    return redirect(url_for('main.index'))

@bp.route('/import-excel', methods=['POST'])
def import_excel():
    try:
        current_app.logger.info('Začátek importu z Excelu')
        if 'excel_file' not in request.files:
            current_app.logger.warning('Nebyl vybrán žádný soubor')
            return jsonify({'error': 'Nebyl vybrán žádný soubor'}), 400
        
        file = request.files['excel_file']
        if file.filename == '':
            current_app.logger.warning('Prázdný název souboru')
            return jsonify({'error': 'Nebyl vybrán žádný soubor'}), 400
        
        if not allowed_file(file.filename):
            current_app.logger.warning(f'Nepovolený typ souboru: {file.filename}')
            return jsonify({'error': 'Nepovolený typ souboru'}), 400
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        df = pd.read_excel(filepath)
        
        # Zpracování dat z Excelu
        df['Server'] = df['Server'].fillna(method='ffill')
        df['Cesta'] = df['Cesta'].fillna(method='ffill')
        df['Název certifikátu'] = df['Název certifikátu'].fillna('Neznámý certifikát')
        
        df = df.dropna(subset=['Server'])
        df = df[df['Expirace'].apply(is_valid_date)]
        
        # Import serverů
        unique_servers = df['Server'].unique()
        for server_name in unique_servers:
            if pd.isna(server_name):
                continue
            if not Server.query.filter_by(nazev=server_name).first():
                new_server = Server(nazev=server_name)
                db.session.add(new_server)
        db.session.commit()
        
        # Import certifikátů
        pridano = aktualizovano = beze_zmeny = 0
        
        for _, row in df.iterrows():
            expirace = datetime.strptime(row['Expirace'], '%d.%m.%Y').date() if isinstance(row['Expirace'], str) else row['Expirace'].date()
            
            existing = Certifikat.query.filter_by(
                server=row['Server'],
                cesta=row['Cesta'],
                nazev=row['Název certifikátu']
            ).first()
            
            if existing:
                if existing.expirace != expirace:
                    existing.expirace = expirace
                    aktualizovano += 1
                else:
                    beze_zmeny += 1
            else:
                new_cert = Certifikat(
                    server=row['Server'],
                    cesta=row['Cesta'],
                    nazev=row['Název certifikátu'],
                    expirace=expirace
                )
                db.session.add(new_cert)
                pridano += 1
                
        db.session.commit()
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'message': f'Import dokončen. Přidáno: {pridano}, Aktualizováno: {aktualizovano}'
        })
        
    except Exception as e:
        current_app.logger.error(f'Chyba při importu: {str(e)}', exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Chyba při importu: {str(e)}'
        }), 400

@bp.route('/export-excel')
def export_excel():
    try:
        certifikaty = Certifikat.query.order_by(Certifikat.server, Certifikat.cesta).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Certifikáty"
        
        # Styly
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        expired_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Hlavička
        headers = ["Server", "Cesta", "Název certifikátu", "Expirace"]
        ws.append(headers)
        
        # Formátování hlavičky
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Data
        today = datetime.now().date()
        row_num = 2
        merge_ranges = {'server': [], 'path': []}
        prev_server = None
        prev_path = None
        merge_start = {'server': 2, 'path': 2}
        
        for cert in certifikaty:
            row = [
                cert.server,
                cert.cesta,
                cert.nazev,
                cert.expirace.strftime('%d.%m.%Y')
            ]
            ws.append(row)
            
            # Formátování řádku
            for cell in ws[row_num]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Alternující barvy řádků
                if row_num % 2 == 0:
                    cell.fill = alt_fill
                
                # Červené podbarvení pro expirované certifikáty - celý řádek
                if cert.expirace < today:
                    cell.fill = expired_fill
            
            # Kontrola pro sloučení buněk serveru
            if prev_server != cert.server and prev_server is not None:
                if row_num - merge_start['server'] > 1:
                    merge_ranges['server'].append(f'A{merge_start["server"]}:A{row_num-1}')
                merge_start['server'] = row_num
            
            # Kontrola pro sloučení buněk cesty
            if prev_path != cert.cesta or prev_server != cert.server:
                if row_num - merge_start['path'] > 1:
                    merge_ranges['path'].append(f'B{merge_start["path"]}:B{row_num-1}')
                merge_start['path'] = row_num
            
            prev_server = cert.server
            prev_path = cert.cesta
            row_num += 1
        
        # Poslední sloučení pro server a cestu
        if row_num - merge_start['server'] > 1:
            merge_ranges['server'].append(f'A{merge_start["server"]}:A{row_num-1}')
        if row_num - merge_start['path'] > 1:
            merge_ranges['path'].append(f'B{merge_start["path"]}:B{row_num-1}')
        
        # Provedení sloučení buněk
        for ranges in merge_ranges.values():
            for cell_range in ranges:
                ws.merge_cells(cell_range)
        
        # Automatická šířka sloupců
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Zamrazení hlavičky
        ws.freeze_panes = "A2"
        
        # Export
        with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            filename = os.path.basename(tmp.name)
            wb.save(tmp.name)
            
            return send_from_directory(
                directory=os.path.dirname(tmp.name),
                path=filename,
                as_attachment=True,
                download_name=f"certifikaty_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
    except Exception as e:
        current_app.logger.error(f'Chyba při exportu do Excelu: {str(e)}')
        flash(f'Chyba při exportu do Excelu: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@bp.route('/smazat-vse')
def smazat_vse():
    try:
        Certifikat.query.delete()
        db.session.commit()
        flash('Všechny certifikáty byly smazány!')
    except Exception as e:
        flash(f'Chyba při mazání: {str(e)}', 'error')
        db.session.rollback()
    return redirect(url_for('main.index'))

@bp.route('/detail/<int:id>')
def detail_certifikatu(id):
    try:
        certifikat = Certifikat.query.get_or_404(id)
        return render_template('detail_modal.html', certifikat=certifikat)
    except Exception as e:
        current_app.logger.error(f'Chyba při načítání detailu: {str(e)}')
        return f'Chyba při načítání detailu: {str(e)}', 500

@bp.route('/get-certificates/<server>')
def get_certificates(server):
    try:
        certifikaty = Certifikat.query.filter_by(server=server)\
            .order_by(Certifikat.expirace, Certifikat.cesta).all()
        
        certs_data = [{
            'id': cert.id,
            'server': cert.server,
            'cesta': cert.cesta,
            'nazev': cert.nazev,
            'expirace': cert.expirace.strftime('%d.%m.%Y'),
            'expiry_class': get_expiry_class(cert)
        } for cert in certifikaty]
        
        return jsonify(certs_data)
    except Exception as e:
        current_app.logger.error(f'Chyba při načítání certifikátů: {str(e)}')
        return jsonify({'error': str(e)}), 500

@bp.route('/evidence_certifikatu/send-report', methods=['POST'])
def trigger_report():
    try:
        from app.tasks import send_monthly_certificate_report
        send_monthly_certificate_report()
        return jsonify({
            'success': True,
            'message': 'Report byl úspěšně odeslán'
        })
    except Exception as e:
        current_app.logger.error(f"Chyba při odesílání reportu: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Chyba při odesílání reportu: {str(e)}'
        }), 500

# ... další routy pro práci s certifikáty ... 